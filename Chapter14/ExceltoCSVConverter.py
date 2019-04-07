#! python3
# Excel to CSV Converter

import openpyxl, os, csv

for excelFile in os.listdir('.'):
    # Skip non-xlsx files, load the workbook object
    if not excelFile.endswith('.xlsx'):
        continue

    wb = openpyxl.load_workbook(excelFile)
    for sheetname in wb.get_sheet_names():
        # Loop through every sheet in the workbook
        sheet = wb.get_sheet_by_name(sheetname)

        # Create the CSV filename from the excel filename and sheet title
        name = excelFile[:-5] + '_' + sheetname + '.csv'
        csvFileObj = open(name, 'w', newline = '')
        #Creat the csv.writer object for this CSV file
        csvWriter = csv.writer(csvFileObj)

        # Loop through every row in the sheet
        for rowNum in range(1, sheet.max_row +1):
            rowData = [] # append each cell to this list
            # Loop through each cell in the row
            for colNum in range(1, sheet.max_column +1):
                # Append each cell's data to rowData
                rowData.append(sheet.cell(row=rowNum, column=colNum).value)

            # Write the rowData list to the CSV file
            for row in rowData:
                csvWriter.writerow(row)

        csvFileObj.close()
