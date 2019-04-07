"""
Create a program blankRowInserter.py that takes two integers and a filename
string as command line arguments. Let’s call the first integer N and the sec-
ond integer M. Starting at row N, the program should insert M blank rows
into the spreadsheet
"""

import openpyxl, sys
from openpyxl.utils import get_column_letter

N = int(sys.argv[1])
M = int(sys.argv[2])
filename = sys.argv[3]

# Load given file
wbIn = openpyxl.load_workbook(filename)
sheetIn = wbIn.get_active_sheet()

# Create output file
wbOut = openpyxl.Workbook()
sheetOut = wbOut.get_sheet_by_name('Sheet')
sheetOut.title = sheetIn.title

for i in range(1,N+1):
    for j in range(1, sheetIn.max_column + 1):
        c = get_column_letter(j)
        sheetOut[c + str(i)] = sheetIn[c + str(i)].value
        #Another way
        #sheetOut.cell(row=i,column=j) = sheetInt(row=i,column=j).value


for i in range(N+M+1, sheetIn.max_row + 1 + M):
    for j in range(1, sheetIn.max_column + 1):
        c = get_column_letter(j)
        sheetOut[c + str(i)] = sheetIn[c + str(i-M)].value

#Save output file
name = "blankRow_" + filename
wbOut.save(name)
