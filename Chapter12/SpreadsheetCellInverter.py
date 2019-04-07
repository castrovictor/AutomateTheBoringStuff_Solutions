"""
Write a program to invert the row and column of the cells in the spread-
sheet. For example, the value at row 5, column 3 will be at row 3, column 5
(and vice versa). This should be done for all cells in the spreadsheet.
"""
import openpyxl, sys
filename = sys.argv[1]


# Load given file
wb = openpyxl.load_workbook(filename)
sheet = wb.get_active_sheet()

# Create new workbook
wbOut = openpyxl.Workbook()
sheetOut = wbOut.get_active_sheet()
sheetOut.title = sheet.title

dataMatrix = []
columns = sheet.max_column
rows = sheet.max_row
print("Max ro, column: " + str(rows) +","+str(columns))

# Read data
for i in range(1,rows+1):
    data = []
    for j in range(1,columns+1):
        data.append(sheet.cell(row=i, column=j).value)
    dataMatrix.append(data)


# Save data inverted
columnIndex = 1
for row in dataMatrix:
    rowIndex = 1
    for cell in row:
        sheetOut.cell(row=rowIndex ,column=columnIndex).value = cell
        rowIndex += 1
    columnIndex += 1


# Save output file
name = "Inverted_" + filename
wbOut.save(name)
