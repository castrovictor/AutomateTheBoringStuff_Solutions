"""Create a program multiplicationTable.py that takes a number N from the com-
mand line and creates an N×N multiplication table in an Excel spreadsheet."""

import openpyxl, sys
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

N = int(sys.argv[1])

#Create sheet
wb = openpyxl.Workbook()
sheet = wb.get_sheet_by_name('Sheet')
sheet.title = ' MultiplicationTable'

# Setting first row and column font style
# Note: Style object is no longer in use, Styles are applied directly to cells
boldFont = Font(bold = True)
#Dividing for loop interations we don't need to check if it's first row or column
upperLimit = N+2

for i in range (2, upperLimit):
    j = get_column_letter(i)
    sheet[j + str(1)].font = boldFont
    sheet['A' + str(i)].font = boldFont
    sheet[j + str(1)] = (i-1)
    sheet['A' + str(i)] = (i-1)

for i in range(2, upperLimit):
    for j in range(2,upperLimit):
        c = get_column_letter(j)
        sheet[c + str(i)] = (i-1) * (j-1)

# Save the Workbook
wb.save('multiplicationTable.xlsx')
